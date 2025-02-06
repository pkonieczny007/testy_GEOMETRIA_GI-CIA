import tkinter as tk
from tkinter import filedialog
import os
import shutil

# Jeśli potrzebujemy obsługi plików XLSX (openpyxl):
try:
    from openpyxl import load_workbook
    OPENPYXL_INSTALLED = True
except ImportError:
    # Jeśli openpyxl nie jest zainstalowane, skrypt w dalszym ciągu może działać bez tej funkcji
    OPENPYXL_INSTALLED = False

# Domyślne wartości
chosen_file = 'dane.txt'                # Domyślny plik z danymi rysunków
chosen_folder = None                    # Folder źródłowy (wybierany przyciskiem)
chosen_destination_folder = None        # Folder docelowy (wybierany przyciskiem)
default_folder_name = 'Pobrane_Pliki'   # Domyślna nazwa folderu docelowego

# ===========================================================
# Funkcje pomocnicze
# ===========================================================

def wczytaj_rysunki(plik):
    """
    Wczytuje listę rysunków z pliku tekstowego.
    Każda linia w pliku to jeden rysunek.
    """
    lista_rysunkow = []
    try:
        with open(plik, 'r', encoding='utf-8') as f:
            for linia in f:
                lista_rysunkow.append(linia.strip())
    except FileNotFoundError:
        pass
    return lista_rysunkow

def przeksztalc_liste(lista):
    """
    Dodatkowe przetwarzanie wczytanej listy. Na razie tylko strip().
    """
    przeksztalcona_lista = []
    for element in lista:
        przeksztalcona_lista.append(element.strip())
    return przeksztalcona_lista

def choose_file():
    """
    Pozwala wybrać plik (np. dane.txt) do wczytania listy rysunków.
    """
    global chosen_file
    file_path = filedialog.askopenfilename(
        initialdir="/",
        title="Wybierz plik z danymi (txt)",
        filetypes=[('Text files', '*.txt'), ('All files', '*.*')]
    )
    if file_path:
        chosen_file = file_path
        result_box.config(state=tk.NORMAL)
        result_box.insert(tk.END, f"Wybrano plik: {chosen_file}\n")
        result_box.config(state=tk.DISABLED)

def choose_folder():
    """
    Pozwala wybrać folder, w którym będziemy szukać rysunków do skopiowania (katalog źródłowy).
    """
    global chosen_folder
    folder_path = filedialog.askdirectory(initialdir="/", title="Wybierz folder z plikami")
    if folder_path:
        chosen_folder = folder_path
        result_box.config(state=tk.NORMAL)
        result_box.insert(tk.END, f"Wybrano katalog źródłowy: {chosen_folder}\n")
        result_box.config(state=tk.DISABLED)

def choose_destination_folder():
    """
    Pozwala wybrać folder docelowy, do którego będą kopiowane pliki.
    """
    global chosen_destination_folder
    folder_path = filedialog.askdirectory(initialdir="/", title="Wybierz folder docelowy")
    if folder_path:
        chosen_destination_folder = folder_path
        destination_folder_entry.delete(0, tk.END)
        destination_folder_entry.insert(0, chosen_destination_folder)

def utworz_plik_dane_z_xlsx():
    """
    Tworzy/uzupełnia plik dane.txt na podstawie pliku XLSX.
    
    - Gdy check „pliki_dld” jest zaznaczony: 
      pobiera wartości z kolumny 'plik_dld' i zapisuje do dane.txt.
    - Gdy check „pliki_dld” nie jest zaznaczony:
      sprawdza kolumnę 'TECHNOLOGIA' i filtruje wg wybranych kryteriów:
         * gięcie  -> 'G' w polu TECHNOLOGIA
         * spawanie -> 'S' w polu TECHNOLOGIA
         * nie spawane -> brak 'S' w polu TECHNOLOGIA
      Zapisuje do dane.txt wartości z kolumny 'Zeinr'.
    """
    if not OPENPYXL_INSTALLED:
        result_box.config(state=tk.NORMAL)
        result_box.insert(tk.END, "Brak biblioteki openpyxl. Zainstaluj ją, aby korzystać z tej funkcji.\n")
        result_box.config(state=tk.DISABLED)
        return

    xlsx_path = filedialog.askopenfilename(
        initialdir="/",
        title="Wybierz plik XLSX",
        filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')]
    )
    if not xlsx_path:
        return

    # Wczytujemy plik excelowy
    wb = load_workbook(xlsx_path)
    sheet = wb.active  # zakładamy, że dane są na pierwszym arkuszu

    # Wczytujemy pierwszy (nagłówkowy) wiersz do listy, by sprawdzić nazwy kolumn
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    # Jeśli checkbox pliki_dld jest zaznaczony, szukamy kolumny plik_dld
    if pliki_dld_var.get():
        try:
            plik_dld_col_idx = header_row.index("plik_dld")
        except ValueError:
            result_box.config(state=tk.NORMAL)
            result_box.insert(tk.END, "Nie znaleziono kolumny 'plik_dld' w pliku.\n")
            result_box.config(state=tk.DISABLED)
            return

        with open("dane.txt", "w", encoding="utf-8") as f:
            # Przechodzimy po wierszach (od 2, bo 1 to nagłówek)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                value_dld = row[plik_dld_col_idx]
                if value_dld:  # jeśli kolumna plik_dld nie jest pusta
                    f.write(str(value_dld).strip() + "\n")

        komunikat = "Plik dane.txt został utworzony z kolumny 'plik_dld'.\n"

    else:
        # Jeśli pliki_dld NIE jest zaznaczone, to działamy wg kolumny TECHNOLOGIA + Zeinr
        try:
            tech_col_idx = header_row.index("TECHNOLOGIA")
            zeinr_col_idx = header_row.index("Zeinr")
        except ValueError:
            # Jeśli nie ma takich kolumn
            result_box.config(state=tk.NORMAL)
            result_box.insert(tk.END, "Nie znaleziono wymaganych kolumn: 'TECHNOLOGIA' i/lub 'Zeinr' w pliku.\n")
            result_box.config(state=tk.DISABLED)
            return

        # Pobieramy aktualne wybrane kryterium z radiobuttona
        kryterium = kryterium_var.get()

        with open("dane.txt", "w", encoding="utf-8") as f:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                technology = row[tech_col_idx]
                zeinr = row[zeinr_col_idx]

                # Upewniamy się, że mamy poprawne wartości w wierszu
                if not technology or not zeinr:
                    continue

                tech_upper = str(technology).upper()

                if kryterium == "giecie":
                    # Szukamy litery G w technologii
                    if 'G' in tech_upper:
                        f.write(str(zeinr).strip() + "\n")

                elif kryterium == "spawanie":
                    # Szukamy litery S w technologii
                    if 'S' in tech_upper:
                        f.write(str(zeinr).strip() + "\n")

                elif kryterium == "nie_spawane":
                    # Szukamy tych, co NIE mają litery S
                    if 'S' not in tech_upper:
                        f.write(str(zeinr).strip() + "\n")

        komunikat = (f"Plik dane.txt został utworzony/uzupełniony na podstawie kolumny TECHNOLOGIA (kryterium: {kryterium}).\n")

    result_box.config(state=tk.NORMAL)
    result_box.insert(tk.END, komunikat)
    result_box.config(state=tk.DISABLED)

def process_list():
    """
    Główna funkcja kopiowania plików:
    1. Pobiera nazwy rysunków z pliku dane.txt (lub innego wybranego).
    2. Wyszukuje je w zadanym folderze (także w podfolderach).
    3. Kopiuje pliki o wybranych rozszerzeniach do folderu docelowego.
    """
    global chosen_file
    global chosen_folder
    global chosen_destination_folder

    # Folder źródłowy – jeśli nie wybrano, ustawiamy na bieżący katalog
    if not chosen_folder:
        folder_zrodlowy = os.getcwd()
    else:
        folder_zrodlowy = chosen_folder

    # Folder docelowy – pobrany z okienka (entry) lub (jeśli puste) domyślnie
    destination_folder = destination_folder_entry.get().strip()
    if not destination_folder:
        destination_folder = default_folder_name
    folder_docelowy = os.path.abspath(destination_folder)

    if not os.path.exists(folder_docelowy):
        os.makedirs(folder_docelowy)

    # Pobieramy wybrane rozszerzenia z checkboxów
    wybrane_rozszerzenia = []
    if pdf_var.get():
        wybrane_rozszerzenia.append(".pdf")
    if dxf_var.get():
        wybrane_rozszerzenia.append(".dxf")
    if dwg_var.get():
        wybrane_rozszerzenia.append(".dwg")
    if tif_var.get():
        wybrane_rozszerzenia.append(".tif")
    if dld_var.get():
        wybrane_rozszerzenia.append(".dld")

    rozszerzenia = tuple(wybrane_rozszerzenia)

    # Wczytujemy rysunki z pliku (np. dane.txt)
    lista_rysunkow = wczytaj_rysunki(chosen_file)
    nazwy_rysunkow = przeksztalc_liste(lista_rysunkow)

    nieskopiowane_rysunki = []
    brakujace_rysunki = []

    # Szukamy plików w folderze źródłowym
    for nazwa_rysunku in nazwy_rysunkow:
        znaleziono_rysunek = False

        for sciezka, foldery, pliki in os.walk(folder_zrodlowy):
            for plik in pliki:
                nazwa, rozszerzenie = os.path.splitext(plik)
                # Sprawdzamy, czy plik ma jedno z wybranych rozszerzeń i zawiera nazwę rysunku
                if rozszerzenie.lower() in rozszerzenia and nazwa_rysunku in nazwa:
                    znaleziono_rysunek = True
                    sciezka_zrodlowa = os.path.join(sciezka, plik)
                    sciezka_docelowa = os.path.join(folder_docelowy, plik)
                    try:
                        shutil.copy(sciezka_zrodlowa, sciezka_docelowa)
                    except shutil.Error:
                        nieskopiowane_rysunki.append(nazwa_rysunku)
                        break

            if znaleziono_rysunek:
                break

        if not znaleziono_rysunek:
            brakujace_rysunki.append(nazwa_rysunku)

    # Wyświetlamy wyniki w polu tekstowym
    result_box.config(state=tk.NORMAL)
    result_box.delete("1.0", tk.END)
    result_box.insert(tk.END, f"Dane pobrane z pliku: {chosen_file}\n")
    result_box.insert(tk.END, f"Ilość nazw rysunków do przetworzenia: {len(nazwy_rysunkow)}\n")
    result_box.insert(tk.END, f"Folder źródłowy: {folder_zrodlowy}\n")
    result_box.insert(tk.END, f"Folder docelowy: {folder_docelowy}\n")
    result_box.insert(tk.END, f"Użyte rozszerzenia: {rozszerzenia}\n\n")

    if nieskopiowane_rysunki:
        result_box.insert(tk.END, "Nie udało się skopiować następujących rysunków:\n")
        result_box.insert(tk.END, "\n".join(nieskopiowane_rysunki) + "\n\n")

    if brakujace_rysunki:
        result_box.insert(tk.END, "Nie znaleziono następujących rysunków:\n")
        result_box.insert(tk.END, "\n".join(brakujace_rysunki) + "\n")

    result_box.config(state=tk.DISABLED)

# ===========================================================
# Konfiguracja głównego okna
# ===========================================================
window = tk.Tk()
window.title("Przetwarzanie listy rysunków - GEOMETRIA GIĘCIA")
window.geometry("620x700")

# Pole tekstowe (instrukcje / opis)
text_box = tk.Label(
    window, 
    text=(
        "Instrukcja:\n"
        "1. Wybierz (opcjonalnie) plik XLSX i utwórz dane.txt z kolumn Zeinr / plik_dld.\n"
        "2. Wybierz (opcjonalnie) plik TXT z listą rysunków (domyślnie dane.txt).\n"
        "3. Wskaż 'Katalog z plikami' (folder źródłowy).\n"
        "4. Ustaw lub wybierz 'Katalog docelowy'.\n"
        "5. Zaznacz wymagane rozszerzenia plików.\n"
        "6. Kliknij 'Przetwórz', aby skopiować znalezione rysunki."
    )
)
text_box.pack(pady=5)

# -------------------------------
# KRYTERIA WYBORU DANYCH Z XLSX
# -------------------------------
kryterium_frame = tk.LabelFrame(window, text="Kryterium generowania pliku dane.txt")
kryterium_frame.pack(padx=5, pady=5, fill="x")

# Radiobuttony - wybór kryterium wg kolumny TECHNOLOGIA:
kryterium_var = tk.StringVar(value="giecie")  # domyślnie "giecie"

radio_giecie = tk.Radiobutton(kryterium_frame, text="gięcie (G, GS, GSO)", variable=kryterium_var, value="giecie")
radio_spawanie = tk.Radiobutton(kryterium_frame, text="spawanie (S, SO, GSO, GS)", variable=kryterium_var, value="spawanie")
radio_nie_spawane = tk.Radiobutton(kryterium_frame, text="nie spawane (bez S)", variable=kryterium_var, value="nie_spawane")

radio_giecie.pack(anchor="w")
radio_spawanie.pack(anchor="w")
radio_nie_spawane.pack(anchor="w")

# Checkbox - czy generować z kolumny 'plik_dld'
pliki_dld_var = tk.BooleanVar(value=False)
pliki_dld_check = tk.Checkbutton(kryterium_frame, text="pliki_dld (z kolumny 'plik_dld' - wynik.xlsx)", variable=pliki_dld_var)
pliki_dld_check.pack(anchor="w", pady=2)

# Przycisk do tworzenia pliku dane.txt z XLSX
xlsx_to_txt_button = tk.Button(kryterium_frame, text="Wczytaj XLSX i utwórz dane.txt", command=utworz_plik_dane_z_xlsx)
xlsx_to_txt_button.pack(pady=5)

# -------------------------------
# WYBÓR PLIKU TXT (opcjonalnie)
# -------------------------------
choose_file_button = tk.Button(window, text="DANE - Wybierz plik txt", command=choose_file)
choose_file_button.pack(pady=5)

# -------------------------------
# WYBÓR KATALOGU ŹRÓDŁOWEGO
# -------------------------------
source_button = tk.Button(window, text="Katalog z plikami", command=choose_folder)
source_button.pack(pady=5)

# -------------------------------
# KATALOG DOCELOWY
# -------------------------------
destination_folder_frame = tk.Frame(window)
destination_folder_frame.pack(pady=5)

destination_label = tk.Label(destination_folder_frame, text="Katalog docelowy:")
destination_label.pack(side=tk.LEFT)

destination_folder_entry = tk.Entry(destination_folder_frame, width=30)
destination_folder_entry.insert(0, default_folder_name)  # domyślnie "Pobrane_Pliki"
destination_folder_entry.pack(side=tk.LEFT, padx=5)

destination_button = tk.Button(destination_folder_frame, text="Wybierz folder", command=choose_destination_folder)
destination_button.pack(side=tk.LEFT)

# -------------------------------
# CHECKBOXY ROZSZERZEŃ
# -------------------------------
extensions_frame = tk.Frame(window)
extensions_frame.pack(pady=5)

pdf_var = tk.BooleanVar(value=False)
dxf_var = tk.BooleanVar(value=True)
dwg_var = tk.BooleanVar(value=False)
tif_var = tk.BooleanVar(value=True)  # Domyślnie zaznaczony
dld_var = tk.BooleanVar(value=False)

tk.Checkbutton(extensions_frame, text=".pdf", variable=pdf_var).pack(side=tk.LEFT)
tk.Checkbutton(extensions_frame, text=".dxf", variable=dxf_var).pack(side=tk.LEFT)
tk.Checkbutton(extensions_frame, text=".dwg", variable=dwg_var).pack(side=tk.LEFT)
tk.Checkbutton(extensions_frame, text=".tif", variable=tif_var).pack(side=tk.LEFT)
tk.Checkbutton(extensions_frame, text=".dld", variable=dld_var).pack(side=tk.LEFT)

# -------------------------------
# PRZYCISK PRZETWARZANIA
# -------------------------------
process_button = tk.Button(window, text="Przetwórz", command=process_list, bg="lightgreen")
process_button.pack(pady=10)

# -------------------------------
# POLE TEKSTOWE Z WYNIKAMI
# -------------------------------
result_box = tk.Text(window, height=15, width=70, state=tk.DISABLED)
result_box.pack(pady=5)

window.mainloop()
