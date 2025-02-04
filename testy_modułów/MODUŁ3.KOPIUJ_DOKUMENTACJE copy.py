import tkinter as tk
from tkinter import filedialog
import os
import shutil

# Jeśli potrzebujemy obsługi plików XLSX (openpyxl):
try:
    from openpyxl import load_workbook
    OPENPYXL_INSTALLED = True
except ImportError:
    # Jeśli openpyxl nie jest zainstalowane, skrypt zadziała bez tej funkcji
    OPENPYXL_INSTALLED = False

# Domyślne wartości
chosen_file = 'dane.txt'          # Domyślny plik z danymi rysunków
default_folder_name = 'Pobrane_Pliki'  # Domyślna nazwa katalogu
chosen_folder = os.getcwd()       # Domyślny folder (jeśli nie użyjemy pola tekstowego)

# ===========================================================
# Funkcje pomocnicze
# ===========================================================

def wczytaj_rysunki(plik):
    """
    Wczytuje listę rysunków z pliku tekstowego.
    Każda linia w pliku to jeden rysunek.
    """
    lista_rysunkow = []
    try:
        with open(plik, 'r', encoding='utf-8') as f:
            for linia in f:
                lista_rysunkow.append(linia.strip())
    except FileNotFoundError:
        pass
    return lista_rysunkow

def przeksztalc_liste(lista):
    """
    Ewentualne dodatkowe przetwarzanie wczytanej listy.
    Na razie tylko strip().
    """
    przeksztalcona_lista = []
    for element in lista:
        przeksztalcona_lista.append(element.strip())
    return przeksztalcona_lista

def choose_file():
    """
    Pozwala wybrać plik (np. dane.txt) do wczytania listy rysunków.
    """
    global chosen_file
    file_path = filedialog.askopenfilename(
        initialdir="/",
        title="Wybierz plik z danymi (txt)",
        filetypes=[('Text files', '*.txt'), ('All files', '*.*')]
    )
    if file_path:
        chosen_file = file_path
        result_box.config(state=tk.NORMAL)
        result_box.insert(tk.END, f"Wybrano plik: {chosen_file}\n")
        result_box.config(state=tk.DISABLED)

def choose_folder():
    """
    Pozwala wybrać folder, w którym będziemy szukać rysunków do skopiowania.
    """
    global chosen_folder
    folder_path = filedialog.askdirectory(initialdir="/", title="Wybierz folder z rysunkami")
    if folder_path:
        chosen_folder = folder_path
        folder_entry.delete(0, tk.END)
        folder_entry.insert(0, chosen_folder)

def utworz_plik_dane_z_xlsx():
    """
    Tworzy plik dane.txt na podstawie pliku wykaz.xlsx.
    Filtruje wiersze, gdzie kolumna 'TECHNOLOGIA' zawiera znak 'G',
    a następnie pobiera wartość z kolumny 'Zeinr' i zapisuje do dane.txt.
    W tym przykładzie zakładamy, że:
      - nazwa kolumny z technologią to 'TECHNOLOGIA'
      - nazwa kolumny z numerem rysunku to 'Zeinr'
    """
    if not OPENPYXL_INSTALLED:
        result_box.config(state=tk.NORMAL)
        result_box.insert(tk.END, "Brak biblioteki openpyxl. Zainstaluj ją, aby korzystać z tej funkcji.\n")
        result_box.config(state=tk.DISABLED)
        return

    xlsx_path = filedialog.askopenfilename(
        initialdir="/",
        title="Wybierz plik XLSX (wykaz.xlsx)",
        filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')]
    )
    if not xlsx_path:
        return

    # Wczytujemy plik excelowy
    wb = load_workbook(xlsx_path)
    sheet = wb.active  # zakładamy, że dane są na pierwszym arkuszu

    # Szukamy indeksów kolumn: TECHNOLOGIA i Zeinr
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        tech_col_idx = header_row.index("TECHNOLOGIA")
        zeinr_col_idx = header_row.index("Zeinr")
    except ValueError:
        # Jeśli nie ma takich kolumn
        result_box.config(state=tk.NORMAL)
        result_box.insert(tk.END, "Nie znaleziono wymaganych kolumn: 'TECHNOLOGIA' i/lub 'Zeinr' w pliku.\n")
        result_box.config(state=tk.DISABLED)
        return

    # Otwieramy (lub tworzymy nowy) plik dane.txt
    with open("dane.txt", "w", encoding="utf-8") as f:
        # Iterujemy od drugiego wiersza (zakładamy, że pierwszy to nagłówki)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            technology = row[tech_col_idx]
            zeinr = row[zeinr_col_idx]

            # Sprawdzamy, czy w polu technologii jest litera 'G'
            if technology and 'G' in technology.upper():  # upper() by uchwycić też np. małe 'g'
                if zeinr:
                    f.write(str(zeinr) + "\n")

    result_box.config(state=tk.NORMAL)
    result_box.insert(tk.END, "Plik dane.txt został utworzony/uzupełniony na podstawie wykazu XLSX.\n")
    result_box.config(state=tk.DISABLED)

def process_list():
    """
    Główna funkcja kopiowania plików:
    1. Pobiera nazwy rysunków z pliku dane.txt (lub innego wybranego).
    2. Wyszukuje je w zadanym folderze (także w podfolderach).
    3. Kopiuje pliki o wybranych rozszerzeniach do folderu docelowego.
    """
    global chosen_file
    global chosen_folder

    # Aktualna zawartość pola tekstowego z nazwą katalogu
    input_folder = folder_entry.get().strip()
    if not input_folder:
        input_folder = default_folder_name

    # Jeśli wpisano ścieżkę względną - przekształcamy na absolutną
    folder_zrodlowy = os.path.abspath(input_folder)
    folder_docelowy = r"C:\PYTHON\PROJEKT_GEOMETRIA_GIĘCIA\testy_modułów\POBRANE"  # można zmienić na inny folder docelowy

    # Tworzymy folder docelowy, jeśli nie istnieje
    if not os.path.exists(folder_docelowy):
        os.makedirs(folder_docelowy)

    # Pobieramy wybrane rozszerzenia z checkboxów
    wybrane_rozszerzenia = []
    if pdf_var.get(): 
        wybrane_rozszerzenia.append(".pdf")
    if dxf_var.get():
        wybrane_rozszerzenia.append(".dxf")
    if dwg_var.get():
        wybrane_rozszerzenia.append(".dwg")
    if tif_var.get():
        wybrane_rozszerzenia.append(".tif")
    if dld_var.get():
        wybrane_rozszerzenia.append(".dld")

    rozszerzenia = tuple(wybrane_rozszerzenia)

    # Wczytujemy rysunki z pliku (np. dane.txt)
    lista_rysunkow = wczytaj_rysunki(chosen_file)
    nazwy_rysunkow = przeksztalc_liste(lista_rysunkow)

    nieskopiowane_rysunki = []
    brakujace_rysunki = []

    # Szukamy plików w folderze źródłowym
    for nazwa_rysunku in nazwy_rysunkow:
        znaleziono_rysunek = False

        # os.walk -> przechodzi przez podkatalogi
        for sciezka, foldery, pliki in os.walk(folder_zrodlowy):
            for plik in pliki:
                nazwa, rozszerzenie = os.path.splitext(plik)
                # Sprawdzamy, czy plik ma jedno z wybranych rozszerzeń i czy w nazwie występuje poszukiwany rysunek
                if rozszerzenie.lower() in rozszerzenia and nazwa_rysunku in nazwa:
                    znaleziono_rysunek = True
                    sciezka_zrodlowa = os.path.join(sciezka, plik)
                    sciezka_docelowa = os.path.join(folder_docelowy, plik)
                    try:
                        shutil.copy(sciezka_zrodlowa, sciezka_docelowa)
                    except shutil.Error:
                        nieskopiowane_rysunki.append(nazwa_rysunku)
                        break

            # Jeśli już znaleziono i skopiowano, nie musimy przeszukiwać dalej
            if znaleziono_rysunek:
                break

        # Po przejściu całego os.walk, jeśli nie znaleziono_rysunek == False, to go brakuje
        if not znaleziono_rysunek:
            brakujace_rysunki.append(nazwa_rysunku)

    # Wyświetlamy wyniki w polu tekstowym
    result_box.config(state=tk.NORMAL)
    result_box.delete("1.0", tk.END)
    result_box.insert(tk.END, f"Dane pobrane z pliku: {chosen_file}\n")
    result_box.insert(tk.END, f"Ilość nazw rysunków do przetworzenia: {len(nazwy_rysunkow)}\n")
    result_box.insert(tk.END, f"Folder źródłowy: {folder_zrodlowy}\n")
    result_box.insert(tk.END, f"Folder docelowy: {folder_docelowy}\n")
    result_box.insert(tk.END, f"Użyte rozszerzenia: {rozszerzenia}\n\n")

    if nieskopiowane_rysunki:
        result_box.insert(tk.END, "Nie udało się skopiować następujących rysunków:\n")
        result_box.insert(tk.END, "\n".join(nieskopiowane_rysunki) + "\n\n")

    if brakujace_rysunki:
        result_box.insert(tk.END, "Nie znaleziono następujących rysunków:\n")
        result_box.insert(tk.END, "\n".join(brakujace_rysunki) + "\n")

    result_box.config(state=tk.DISABLED)

# ===========================================================
# Konfiguracja głównego okna
# ===========================================================
window = tk.Tk()
window.title("Przetwarzanie listy rysunków - GEOMETRIA GIĘCIA")
window.geometry("600x600")

# Pole tekstowe (opcjonalne) do wyświetlania komunikatów / instrukcji
text_box = tk.Label(window, text=(
    "Instrukcja:\n"
    "1. (Opcjonalne) Kliknij 'Wczytaj XLSX i utwórz dane.txt', aby wygenerować plik z nazwami rysunków.\n"
    "2. (Opcjonalne) Wybierz inny plik tekstowy z listą rysunków (domyślnie dane.txt).\n"
    "3. Wpisz nazwę folderu z plikami (domyślnie Pobrane_Pliki) lub wybierz przyciskiem.\n"
    "4. Zaznacz wymagane rozszerzenia plików.\n"
    "5. Kliknij 'Przetwórz', aby skopiować znalezione rysunki."
))
text_box.pack(pady=5)

# Przycisk do tworzenia pliku dane.txt z wykaz.xlsx
xlsx_to_txt_button = tk.Button(window, text="Wczytaj XLSX i utwórz dane.txt", command=utworz_plik_dane_z_xlsx)
xlsx_to_txt_button.pack(pady=2)

# Przycisk do wybrania pliku txt
choose_file_button = tk.Button(window, text="DANE - Wybierz plik txt", command=choose_file)
choose_file_button.pack(pady=2)

# Pole do wpisania folderu (domyślnie Pobrane_Pliki)
folder_frame = tk.Frame(window)
folder_frame.pack(pady=2)

folder_label = tk.Label(folder_frame, text="Katalog z plikami:")
folder_label.pack(side=tk.LEFT)

folder_entry = tk.Entry(folder_frame, width=30)
folder_entry.insert(0, default_folder_name)  # domyślnie "Pobrane_Pliki"
folder_entry.pack(side=tk.LEFT, padx=5)

folder_button = tk.Button(folder_frame, text="Wybierz folder", command=choose_folder)
folder_button.pack(side=tk.LEFT)

# Checkboxy rozszerzeń
extensions_frame = tk.Frame(window)
extensions_frame.pack(pady=5)

pdf_var = tk.BooleanVar(value=False)
dxf_var = tk.BooleanVar(value=True)
dwg_var = tk.BooleanVar(value=False)
tif_var = tk.BooleanVar(value=False)
dld_var = tk.BooleanVar(value=False)

tk.Checkbutton(extensions_frame, text=".pdf", variable=pdf_var).pack(side=tk.LEFT)
tk.Checkbutton(extensions_frame, text=".dxf", variable=dxf_var).pack(side=tk.LEFT)
tk.Checkbutton(extensions_frame, text=".dwg", variable=dwg_var).pack(side=tk.LEFT)
tk.Checkbutton(extensions_frame, text=".tif", variable=tif_var).pack(side=tk.LEFT)
tk.Checkbutton(extensions_frame, text=".dld", variable=dld_var).pack(side=tk.LEFT)

# Przycisk przetwarzania
process_button = tk.Button(window, text="Przetwórz", command=process_list, bg="lightgreen")
process_button.pack(pady=5)

# Pole tekstowe do wyświetlania wyników
result_box = tk.Text(window, height=15, width=70, state=tk.DISABLED)
result_box.pack(pady=5)

window.mainloop()
